# Excel Editor
# pip install openpyxl
import openpyxl as excel
from openpyxl.styles import Font
# load Excel File
wbook = excel.load_workbook('test.xlsx')
ws = wbook.active
# read all data
for row in ws.rows:
    for cell in row:
        print(cell.value, end=' ')
    print()
# read data by cell
print(ws['A1'].value)
# read data by row
for row in ws.rows:
    print(row[0].value, row[1].value, row[2].value)
# read data by column
for col in ws.columns:
    print(col[0].value, col[1].value, col[2].value)
# write data
wbook = excel.Workbook()
ws = wbook.active
# write data
ws['A1'] = 'Hello'
# write data to specific cell
ws.cell(row=1, column=2).value = 'World'
# append data
ws.append(['Python', 'is', 'good'])
# change font
ws['A1'].font = Font(size=20, color='FF0000')
# change column width
ws.column_dimensions['A'].width = 30
# change row height
ws.row_dimensions[1].height = 50
# Delete row
ws.delete_rows(1)
# Delete column
ws.delete_cols(1)
# Delete Cell
ws['A1'] = ''
# Save Excel File
wbook.save('test.xlsx')